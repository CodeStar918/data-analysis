"""测试工具：用 openpyxl 构造内存中的 Excel 文件。"""
import io
from datetime import datetime

from openpyxl import Workbook


def build_xlsx(sheets: dict, merges: dict | None = None) -> bytes:
    """sheets: {sheet名: [[行], ...]}; merges: {sheet名: ["A1:B2", ...]}"""
    wb = Workbook()
    first_used = False
    for name, rows in sheets.items():
        if not first_used:
            ws = wb.active
            ws.title = name
            first_used = True
        else:
            ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
        for m in (merges or {}).get(name, []):
            ws.merge_cells(m)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


SAMPLE_TIME = datetime(2024, 6, 1, 10, 30, 0)
